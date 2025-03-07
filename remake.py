# OPENS 'output.xlsx' REMAKES:
#   1. MWF Labs
#   2. TR Labs
#   3. Instructors

# Library Files
import xlrd # NEED xlrd TO READ FROM .xlsx FILE
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
import pandas as pd

# Helper Function Files
import daily, helper, compiled

# MAIN FUNCTION
def remake_from_list_view():
    # Declare Input/Output Files
    input_file = 'output.xlsx'
    output_file = 'finalOutput.xlsx'
    
    # Convert "List View" in 'output.xlsx' into a list of dictionaries
    # This list WILL contain conflicts
    adjusted_list = xlsx_to_dicts(input_file)

    test_person_conflict = [
        {'course': 'CPE 464', 'section': 1, 'instructor': 'Bellardo, John', 'type': 'Lecture', 'room': '052-0E26', 'days': 'TR', 'start': '9:10am', 'end': '11am', 'note': None, 'id': None},
        {'course': 'CPE 464', 'section': 2, 'instructor': 'Bellardo, John', 'type': 'Laboratory', 'room': '020-0124', 'days': 'TR', 'start': '9:10am', 'end': '12pm', 'note': None, 'id': None},
        {'course': 'CPE 464', 'section': 3, 'instructor': 'Staff', 'type': 'Lecture', 'room': '010-0200', 'days': 'TR', 'start': '12:10pm', 'end': '1:30pm', 'note': None, 'id': None},
        {'course': 'CPE 464', 'section': 4, 'instructor': 'Staff', 'type': 'Laboratory', 'room': '020-0124', 'days': 'W', 'start': '3:10pm', 'end': '6pm', 'note': None, 'id': None},
        {'course': 'CPE 523', 'section': 1, 'instructor': 'Danowitz, Andrew', 'type': 'Seminar', 'room': '020-0100', 'days': 'TR', 'start': '8:10am', 'end': '9:30am', 'note': None, 'id': None},
        {'course': 'CPE 523', 'section': 2, 'instructor': 'Danowitz, Andrew', 'type': 'Laboratory', 'room': '020-0100', 'days': 'TR', 'start': '9:40am', 'end': '11am', 'note': None, 'id': None}
    ]

    test_location_conflict = [
        {'course': 'CPE 464', 'section': 1, 'instructor': 'Bellardo, John', 'type': 'Lecture', 'room': '020-0100', 'days': 'TR', 'start': '9:40am', 'end': '11am', 'note': None, 'id': None},
        {'course': 'CPE 464', 'section': 2, 'instructor': 'Bellardo, John', 'type': 'Laboratory', 'room': '020-0124', 'days': 'W', 'start': '9:10am', 'end': '12pm', 'note': None, 'id': None},
        {'course': 'CPE 464', 'section': 3, 'instructor': 'Staff', 'type': 'Lecture', 'room': '010-0200', 'days': 'TR', 'start': '12:10pm', 'end': '1:30pm', 'note': None, 'id': None},
        {'course': 'CPE 464', 'section': 4, 'instructor': 'Staff', 'type': 'Laboratory', 'room': '020-0124', 'days': 'W', 'start': '3:10pm', 'end': '6pm', 'note': None, 'id': None},
        {'course': 'CPE 523', 'section': 1, 'instructor': 'Danowitz, Andrew', 'type': 'Seminar', 'room': '020-0100', 'days': 'TR', 'start': '8:10am', 'end': '9:30am', 'note': None, 'id': None},
        {'course': 'CPE 523', 'section': 2, 'instructor': 'Danowitz, Andrew', 'type': 'Laboratory', 'room': '020-0100', 'days': 'TR', 'start': '9:40am', 'end': '11am', 'note': None, 'id': None}
    ]

    test_conflicts = [
        {'course': 'CPE 1', 'section': 1, 'instructor': 'Bellardo, John', 'type': 'Lecture', 'room': '052-0E26', 'days': 'TR', 'start': '9:10am', 'end': '11am', 'note': None, 'id': None},
        {'course': 'CPE 2', 'section': 2, 'instructor': 'Bellardo, John', 'type': 'Laboratory', 'room': '020-0124', 'days': 'TR', 'start': '9:10am', 'end': '12pm', 'note': None, 'id': None},
        {'course': 'CPE 3', 'section': 3, 'instructor': 'Staff', 'type': 'Lecture', 'room': '010-0200', 'days': 'TR', 'start': '12:10pm', 'end': '1:30pm', 'note': None, 'id': None},
        {'course': 'CPE 4', 'section': 4, 'instructor': 'Staff', 'type': 'Laboratory', 'room': '020-0124', 'days': 'W', 'start': '3:10pm', 'end': '6pm', 'note': None, 'id': None},
        {'course': 'CPE 5', 'section': 1, 'instructor': 'Danowitz, Andrew', 'type': 'Seminar', 'room': '020-0100', 'days': 'TR', 'start': '8:10am', 'end': '9:30am', 'note': None, 'id': None},
        {'course': 'CPE 6', 'section': 2, 'instructor': 'Danowitz, Andrew', 'type': 'Laboratory', 'room': '020-0100', 'days': 'TR', 'start': '9:40am', 'end': '11am', 'note': None, 'id': None},
        {'course': 'CPE 7', 'section': 1, 'instructor': 'Bellardo, John', 'type': 'Lecture', 'room': '020-0100', 'days': 'TR', 'start': '9:40am', 'end': '11am', 'note': None, 'id': None},
        {'course': 'CPE 8', 'section': 2, 'instructor': 'Bellardo, John', 'type': 'Laboratory', 'room': '020-0124', 'days': 'W', 'start': '9:10am', 'end': '12pm', 'note': None, 'id': None},
        {'course': 'CPE 9', 'section': 3, 'instructor': 'Staff', 'type': 'Lecture', 'room': '010-0200', 'days': 'TR', 'start': '12:10pm', 'end': '1:30pm', 'note': None, 'id': None},
        {'course': 'CPE 10', 'section': 4, 'instructor': 'Staff', 'type': 'Laboratory', 'room': '020-0124', 'days': 'W', 'start': '3:10pm', 'end': '6pm', 'note': None, 'id': None},
        {'course': 'CPE 11', 'section': 1, 'instructor': 'Danowitz, Andrew', 'type': 'Seminar', 'room': '020-0100', 'days': 'TR', 'start': '8:10am', 'end': '9:30am', 'note': None, 'id': None},
        {'course': 'CPE 12', 'section': 2, 'instructor': 'Danowitz, Andrew', 'type': 'Laboratory', 'room': '020-0100', 'days': 'TR', 'start': '9:40am', 'end': '11am', 'note': None, 'id': None}

    ]

    # Check conflicts
        # Place Conflict --> 1 person, 2 places, 1 time
            # Professor is in 2 places at once
        # Person Conflict --> 2 people, 1 place, 1 time
            # Two professors in the same place at once
    
    person_conflict, location_conflict = generate_conflict_lists(adjusted_list)    

    # Generate a list of dicts that have NO conflicts
    valid_list = generate_valid_list(adjusted_list, person_conflict, location_conflict)

    # Make new excel File
    wb = openpyxl.Workbook()

    # Rename the default sheet to 'MWF Labs'
    ws1 = wb.active

    ws1.title = 'MWF Labs'
    ws2 = wb.create_sheet('TR Labs')
    ws3 = wb.create_sheet('Intructors')
    ws4 = wb.create_sheet('List View')
    

    # ALL unique Rooms
    all_unique_rooms= get_sorted_unique_rooms(adjusted_list)

    # MWF unique rooms
    MWF_unique_rooms = day_unique_rooms(adjusted_list, 'MWF')

    # TR unique rooms
    TR_unique_rooms = day_unique_rooms(adjusted_list, 'TR')

    # ALL unique Instructors (SORTED)
    all_unique_intructors = unique_instructors(adjusted_list)


    # Make Each Sheet using the adjusted_list of Dictionaries
    remake_MWF(ws1, valid_list, MWF_unique_rooms)
    remake_TR(ws2, valid_list, TR_unique_rooms)
    remake_Intructors(ws3, valid_list, all_unique_intructors)
    remake_ListView(ws4, valid_list, person_conflict, location_conflict)

    # Save the workbook to a file
    wb.save(output_file)


# Generates two lists:
#   1. Place Conflict --> 1 person, 2 places, 1 time
#       - Professor is in 2 places at once
#   2. Person Conflict --> 2 people, 1 place, 1 time
#       - Two professors in the same place at once
def generate_conflict_lists(data):
    # Initialize conflict lists
    place_conflict = []
    person_conflict = []

    # Track combinations to identify conflicts
    instructor_tracker = {}
    room_tracker = {}

    # Iterate over each record
    for record in data:
        # Extract relevant fields
        days = record.get('days')
        start = record.get('start')
        instructor = record.get('instructor')
        room = record.get('room')

        # Skip records with missing critical information
        if not (days and start and instructor and room) or instructor == 'Staff':
            continue

        # Check for conflicts in instructor scheduling
        instructor_key = (days, start, instructor)
        if instructor_key not in instructor_tracker:
            instructor_tracker[instructor_key] = [record]
        else:
            # Check for room conflict
            for existing in instructor_tracker[instructor_key]:
                if existing['room'] != room:
                    if record not in place_conflict:
                        place_conflict.append(record)
                    if existing not in place_conflict:
                        place_conflict.append(existing)
            instructor_tracker[instructor_key].append(record)

        # Check for conflicts in room scheduling
        room_key = (days, start, room)
        if room_key not in room_tracker:
            room_tracker[room_key] = [record]
        else:
            # Check for instructor conflict
            for existing in room_tracker[room_key]:
                if existing['instructor'] != instructor:
                    if record not in person_conflict:
                        person_conflict.append(record)
                    if existing not in person_conflict:
                        person_conflict.append(existing)
            room_tracker[room_key].append(record)

    return place_conflict, person_conflict

# Generate a list of VALID dictionaries
# VALID: The dictionary representing a course has no
#        scheduling conflicts.
def generate_valid_list(adjusted_list, person_conflict, place_conflict):
    # Create an empty list to hold the valid dictionaries
    valid_list = []

    # Combine person_conflict and place_conflict into a single list
    combined_conflicts = person_conflict + place_conflict

    # Loop through each dictionary in adjusted_list
    for adjusted_item in adjusted_list:
        # Assume the item is valid initially
        is_valid = True

        # Check if the dictionary exists in the combined_conflicts list
        for conflict_item in combined_conflicts:
            # Compare the dictionaries directly
            if adjusted_item == conflict_item:
                # If a conflict is found, mark it as invalid and break out of the loop
                is_valid = False
                break

        # If the dictionary is valid, add it to the valid_list
        if is_valid:
            valid_list.append(adjusted_item)

    # Return the valid_list
    return valid_list

# Remake MWF Labs Sheet using List[Dictionaries]
def remake_MWF(sheet, valid_list, unique_rooms):
    # Add the "Time Column" as the left most column
    compiled.add_time_column(sheet)
    compiled.add_columns(sheet, unique_rooms)

    for record in valid_list:
        room = record['room']
        start_time = helper.round_class_time_helper(record['start'])
        end_time = helper.round_class_time_helper(record['end'])
        instructor = record['instructor']
        course = record['course']
        section = record['section']
        #num_cells = helper.find_number_of_cells(start_time, end_time)
        num_cells = compiled.calculate_chunks(start_time, end_time)
        class_type = record['type']
        days = record['days']
        remake_schedule_MWF(sheet, unique_rooms, room, start_time, end_time, instructor, course, section, num_cells, class_type, days)

    # Adjust Width of Columns
    compiled.adjust_column_widths(sheet)

# Remake TR Labs Sheet using List[Dictionaries]
def remake_TR(sheet, valid_list, unique_rooms):
    # Add the "Time Column" as the left most column
    compiled.add_time_column(sheet)
    compiled.add_columns(sheet, unique_rooms)


    for record in valid_list:
        room = record['room']
        start_time = helper.round_class_time_helper(record['start'])
        end_time = helper.round_class_time_helper(record['end'])
        instructor = record['instructor']
        course = record['course']
        section = record['section']
        #num_cells = helper.find_number_of_cells(start_time, end_time)
        num_cells = compiled.calculate_chunks(start_time, end_time)
        class_type = record['type']
        days = record['days']
        remake_schedule_TR(sheet, unique_rooms, room, start_time, end_time, instructor, course, section, num_cells, class_type, days)

    # Adjust Width of Columns
    compiled.adjust_column_widths(sheet)
    ...

# Remake Instructors Sheet using List[Dictionaries]
def remake_Intructors(sheet, valid_list, all_unique_intructors):
    # Add the "Time Column" as the left most column
    compiled.add_time_column(sheet)
    compiled.add_columns_instructors(sheet, all_unique_intructors)

    for record in valid_list:
        room = record['room']
        start_time = helper.round_class_time_helper(record['start'])
        end_time = helper.round_class_time_helper(record['end'])
        instructor = record['instructor']
        course = record['course']
        section = record['section']
        #num_cells = helper.find_number_of_cells(start_time, end_time)
        num_cells = compiled.calculate_chunks(start_time, end_time)
        class_type = record['type']
        days = record['days']

        remake_schedule_instructor(sheet, all_unique_intructors, room, start_time, end_time, instructor, course, section, num_cells, class_type, days)

    # Adjust Width of Columns
    compiled.adjust_column_widths(sheet)
    ...

# Remake List View Sheet using List[Dictionaries]
# List View will contain any new time conflicts after adjustment
def remake_ListView(sheet, valid_list, person_conflicts, place_conflicts):
    compiled.add_columns_list_view(sheet)

    # Initialize the row index for writing to the sheet
    current_row_index = 2

    # Convert conflict lists to sets for fast membership checking
    person_conflicts_set = set(tuple(d.items()) for d in person_conflicts)
    place_conflicts_set = set(tuple(d.items()) for d in place_conflicts)

    # Loop over the list of dictionaries and write non-conflicting data to the sheet
    for record in valid_list:
        # Skip entries that are in either conflict list
        if (tuple(record.items()) in person_conflicts_set or
                tuple(record.items()) in place_conflicts_set):
            continue

        # Write Course value to the first column
        sheet.cell(row=current_row_index, column=1, value=record.get('course'))

        # Write Section value to the second column
        sheet.cell(row=current_row_index, column=2, value=record.get('section'))

        # Write Instructor value to the third column
        sheet.cell(row=current_row_index, column=3, value=record.get('instructor'))

        # Write Type value to the fourth column
        sheet.cell(row=current_row_index, column=4, value=record.get('type'))

        # Write Room value to the fifth column
        sheet.cell(row=current_row_index, column=5, value=record.get('room'))

        # Write Days value to the sixth column
        sheet.cell(row=current_row_index, column=6, value=record.get('days'))

        # Write Start Time value to the seventh column
        sheet.cell(row=current_row_index, column=7, value=record.get('start'))

        # Write End Time value to the eighth column
        sheet.cell(row=current_row_index, column=8, value=record.get('end'))

        # Increment the row index after writing valid data
        current_row_index += 1

    # Skip a row and write "Place Conflicts"
    current_row_index += 1
    sheet.cell(row=current_row_index, column=1, value="Place Conflicts")
    current_row_index += 1

    # Write all place conflict dictionaries
    for conflict in place_conflicts:
        sheet.cell(row=current_row_index, column=1, value=conflict.get('course'))
        sheet.cell(row=current_row_index, column=2, value=conflict.get('section'))
        sheet.cell(row=current_row_index, column=3, value=conflict.get('instructor'))
        sheet.cell(row=current_row_index, column=4, value=conflict.get('type'))
        sheet.cell(row=current_row_index, column=5, value=conflict.get('room'))
        sheet.cell(row=current_row_index, column=6, value=conflict.get('days'))
        sheet.cell(row=current_row_index, column=7, value=conflict.get('start'))
        sheet.cell(row=current_row_index, column=8, value=conflict.get('end'))
        current_row_index += 1

    # Skip a row and write "Person Conflicts"
    current_row_index += 1
    sheet.cell(row=current_row_index, column=1, value="Person Conflicts")
    current_row_index += 1

    # Write all person conflict dictionaries
    for conflict in person_conflicts:
        sheet.cell(row=current_row_index, column=1, value=conflict.get('course'))
        sheet.cell(row=current_row_index, column=2, value=conflict.get('section'))
        sheet.cell(row=current_row_index, column=3, value=conflict.get('instructor'))
        sheet.cell(row=current_row_index, column=4, value=conflict.get('type'))
        sheet.cell(row=current_row_index, column=5, value=conflict.get('room'))
        sheet.cell(row=current_row_index, column=6, value=conflict.get('days'))
        sheet.cell(row=current_row_index, column=7, value=conflict.get('start'))
        sheet.cell(row=current_row_index, column=8, value=conflict.get('end'))
        current_row_index += 1

    # Adjust Width of Columns
    compiled.adjust_column_widths_LV(sheet)
    ...

# Fills the appropriate cell with the proper format given the series of inputs
# in the 'Instructor' Sheet
def remake_schedule_instructor(sheet, unique_instructors, lab_room, start_time, end_time, instructor, course, section, num_cells, component, days):
    # Find the starting row based on the start time
    start_row = None
    end_row = None
    for row in range(2, sheet.max_row + 1):
        time_value = sheet.cell(row=row, column=1).value
        if time_value == start_time:
            start_row = row
            break

    if start_row is None:
        print(f"Error: Start time {start_time} not found in the 'Time' column.")
        return

    # Determine the base column for the instructor
    try:
        instructor_index = unique_instructors.index(instructor)
        base_column = 2 + instructor_index * 2  # Each instructor gets two columns
    except ValueError:
        print(f"Error: Instructor {instructor} not found in unique_instructors.")
        return

    # Assign to the appropriate column based on days
    if any(day in days for day in ['M', 'W', 'F']):
        column = base_column  # First column for the instructor
    else:
        column = base_column + 1  # Second column for the instructor

    # Format the class information
    if lab_room == 'University Lecture Room':
        formatted_value = f"{instructor}\n{course}-{int(section)}\n{component}\n{days} {start_time}-{end_time}\n{lab_room}"
    else:
        formatted_value = f"{instructor}\n{course}-{int(section)}\n{component}\n{days} {start_time}-{end_time}\nRoom {lab_room}"

    # Populate the cells with the class information
    cell = sheet.cell(row=start_row, column=column)
    cell.value = formatted_value
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    merge_row_cells(sheet, start_row, num_cells, start_time, column)
    
# Fills the appropriate cell with the proper format given the series of inputs
# in the 'MWF Labs' Sheet
def remake_schedule_MWF(sheet, unique_rooms, lab_room, start_time, end_time, instructor, course, section, num_cells, component, days):
    # Skip entries that don't meet the required conditions
    if not ('M' in days or 'W' in days or 'F' in days) or component != 'Laboratory':
        return

    # Find the starting row based on the start time
    start_row = None
    for row in range(2, sheet.max_row + 1):
        time_value = sheet.cell(row=row, column=1).value
        if time_value == start_time:
            start_row = row
            break

    if start_row is None:
        print(f"Error: Start time {start_time} not found in the 'Time' column.")
        return

    # Determine the column for the room
    try:
        room_index = unique_rooms.index(lab_room)
        column = 2 + room_index  # Each room gets one column
    except ValueError:
        print(f"Error: Room {lab_room} not found in unique_rooms. {instructor}, {course}, {days}")
        return

    # Format the class information
    formatted_value = f"{instructor}\n{course}-{int(section)}\n{component}\n{days} {start_time}-{end_time}"

    # Populate the cell with the class information
    cell = sheet.cell(row=start_row, column=column)
    cell.value = formatted_value
    cell.alignment = Alignment(wrap_text=True, vertical='top')

    merge_row_cells(sheet, start_row, num_cells, start_time, column)


# Fills the appropriate cell with the proper format given the series of inputs
# in the 'TR Labs' Sheet
def remake_schedule_TR(sheet, unique_rooms, lab_room, start_time, end_time, instructor, course, section, num_cells, component, days):
    # Skip entries that don't meet the required conditions
    if not ('T' in days or 'R' in days) or component != 'Laboratory':
        return

    # Find the starting row based on the start time
    start_row = None
    for row in range(2, sheet.max_row + 1):
        time_value = sheet.cell(row=row, column=1).value
        if time_value == start_time:
            start_row = row
            break

    if start_row is None:
        print(f"Error: Start time {start_time} not found in the 'Time' column.")
        return

    # Determine the column for the room
    try:
        room_index = unique_rooms.index(lab_room)
        column = 2 + room_index  # Each room gets one column
    except ValueError:
        print(f"Error: Room {lab_room} not found in unique_rooms. {instructor}, {course}, {days}")
        return

    # Format the class information
    formatted_value = f"{instructor}\n{course}-{int(section)}\n{component}\n{days} {start_time}-{end_time}"

    # Populate the cell with the class information
    cell = sheet.cell(row=start_row, column=column)
    cell.value = formatted_value
    cell.alignment = Alignment(wrap_text=True, vertical='top')

    merge_row_cells(sheet, start_row, num_cells, start_time, column)


def merge_row_cells(sheet, start_row, num_cells, start_time, column):
    # Merge cells based on the num_cells
    end_row = start_row + num_cells - 1
    if end_row > sheet.max_row:
        print(f"Error: Cannot merge cells beyond the last row for start time {start_time}.")
        return

    sheet.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)

    # Format the merged cell
    merged_cell = sheet.cell(row=start_row, column=column)
    merged_cell.alignment = Alignment(wrap_text=True, vertical='top')


# Takes xlsx file --> list of dictionaries
def xlsx_to_dicts(input_file):
    xlsx_file = input_file

    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb.worksheets[3]  # Accessing the 4th sheet, "List View" (0-based index)

    # Read the first row as keys
    keys = [cell.value for cell in sheet[1]]

    # Read the remaining rows as dictionaries
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {keys[i]: row[i] for i in range(len(keys))}
        data.append(row_dict)

    return data

# List of Dicts --> List of Unique instructors
def unique_instructors(data):
    # Initialize an empty set to store unique instructor values
    unique_instructors = set()

    # Iterate over each dictionary in the list
    for record in data:
        # Check if 'instructor' key exists and has a valid value
        if 'instructor' in record and record['instructor']:
            unique_instructors.add(record['instructor'])

    # Convert the set to a sorted list
    sorted_instructors = sorted(unique_instructors)

    # Return the sorted list of instructors
    return sorted_instructors

# List of Dicts --> List of ALL Unique rooms
def get_sorted_unique_rooms(data):
    # Initialize an empty set to store unique room values
    unique_rooms = set()

    # Iterate over each dictionary in the list
    for record in data:
        # Check if the 'room' key exists and its value is not None or empty
        if 'room' in record and record['room']:
            # Add the room value to the set
            unique_rooms.add(record['room'])

    # Convert the set to a sorted list
    sorted_rooms = sorted(unique_rooms)

    # Return the sorted list of rooms
    return sorted_rooms

# List of Dicts --> List of Unique Rooms Matching MWF or TR Days and Type == Laboratory
def day_unique_rooms(data, days_group):
    # Initialize an empty set to store unique room values
    unique_rooms = set()

    # Define the allowed day groups
    valid_days = {'MWF': {'M', 'W', 'F'}, 'TR': {'T', 'R'}}

    # Ensure the input days_group is valid
    if days_group not in valid_days:
        raise ValueError("Invalid days input. Use 'MWF' or 'TR'.")

    # Get the set of valid days for the input group
    required_days = valid_days[days_group]

    # Iterate over each dictionary in the list
    for record in data:
        # Check if 'days' key exists and overlaps with the required days
        record_days = set(record.get('days', ''))  # Get 'days' as a set of characters
        if (
            required_days.intersection(record_days)  # Days overlap with required days
            and 'room' in record and record['room']  # Room key exists and is valid
            and record.get('type') == 'Laboratory'  # Type is 'Laboratory'
        ):
            # Add the room value to the set
            unique_rooms.add(record['room'])

    # Convert the set to a sorted list
    sorted_rooms = sorted(unique_rooms)

    # Return the sorted list of rooms
    return sorted_rooms

# Prints each Dictionary from "List View" and Number of Classes Scheduled
def test_out(list_of_dicts):
    for item in list_of_dicts:
        print(item)
    print(len(list_of_dicts))

if __name__ == "__main__":
    remake_from_list_view()
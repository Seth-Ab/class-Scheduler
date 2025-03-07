# MAIN FILE

# IN THIS FILE:
# 1. Access finalSheet.csv
# 2. Extract Information
# 3. Generate new excel files (2242_schedules.xls): 
#   - MWF Labs
#   - TR Labs
#   - Instructors
#   - List View

# Library Files
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
import pandas as pd

# Helper Function Files
import daily, helper, remake

def main():
    input_file = 'sheetFinal.csv'
    output_file = 'output.xlsx'

    # List of Dictionaries of INPUTS
    all_data = csv_to_dicts(input_file)

    # Make a new list of Directories where
    # Key 'Component' --> Value: 'Laboratory'
    lab_data = [entry for entry in all_data if entry.get('Component') == 'Laboratory'] # 28
    # Key 'Component' --> Value: 'Lecture'
    lecture_data = [entry for entry in all_data if entry.get('Component') == 'Lecture'] # 29
    # Key 'Component' --> Value: 'Laboratory' OR 'Lecture'
    lab_and_lecture_data = [entry for entry in all_data if (entry.get('Component') == 'Laboratory') or (entry.get('Component') == 'Lecture')] # 57

    lec_lab_sem = [entry for entry in all_data if (entry.get('Component') == 'Laboratory') or (entry.get('Component') == 'Lecture') or (entry.get('Component') == 'Seminar')]

    # UNIQUE Rooms
    unique_rooms= helper.get_unique_rooms(lab_data)
    # UNIQUE Instructors
    unique_instructors = helper.get_unique_instructors(lab_data)

    unique_instructors_2 = helper.get_unique_instructors(lab_and_lecture_data)
    # Make new excel File
    wb = Workbook()

    # Rename the default sheet to 'MWF Labs'
    ws1 = wb.active

    ws1.title = 'MWF Labs'
    ws2 = wb.create_sheet('TR Labs')
    ws3 = wb.create_sheet('Intructors')
    ws4 = wb.create_sheet('List View')

    # Make Each Sheet
    daily.make_mwf_labs(ws1, unique_rooms, lab_data)
    daily.make_tr_labs(ws2, unique_rooms, lab_data)
    daily.make_instructors(ws3, unique_instructors_2, lec_lab_sem)
    daily.make_lab_view(ws4, all_data)

    # Save the workbook to a file
    wb.save(output_file)

# Adjusts the width of columns in the scheduling Sheets
def adjust_column_widths(worksheet):
    # Iterate over all columns in the worksheet
    for col in worksheet.columns:
        column = col[0].column_letter  # Get the column name
        worksheet.column_dimensions[column].width = 25


# Adjusts the width of columns in the 'List View' Sheet
def adjust_column_widths_LV(worksheet):
    # Iterate over all columns in the worksheet
    for col in worksheet.columns:
        column = col[0].column_letter  # Get the column name
        worksheet.column_dimensions[column].width = 13


# Adds the Rooms to the Sheet
def add_columns(worksheet, column_titles):
    # Add column titles to the top row
    for index, title in enumerate(column_titles, start=2):
        worksheet.cell(row=1, column=index, value=title)


# Add Instructors column in the Intructor sheet
def add_columns_instructors(worksheet, column_titles):
    # Add column titles to the top row
    column_index = 2  # Starting from the second column
    for title in column_titles:
        worksheet.cell(row=1, column=column_index, value=f"{title} MWF")
        worksheet.cell(row=1, column=column_index + 1, value=f"{title} TR")
        column_index += 2  # Move to the next pair of columns

# Add List View Sections
def add_columns_list_view(worksheet):
    column_titles = ['course','section','instructor','type','room','days','start','end','note','id']
    for col_num, title in enumerate(column_titles, start=1):
        worksheet.cell(row=1, column=col_num, value=title)



# Adds a time column to a worksheet
def add_time_column(worksheet):
    # Add 'Time' label in the first cell
    worksheet['A1'] = 'Time'

    # Initialize start time and end time
    start_time = datetime.strptime("07:00 AM", "%I:%M %p")
    end_time = datetime.strptime("10:00 PM", "%I:%M %p")
    
    # Generate times in 30-minute increments
    current_time = start_time
    row_index = 2  # Start populating from the second row
    while current_time <= end_time:
        worksheet.cell(row=row_index, column=1, value=current_time.strftime("%I:%M %p"))
        current_time += timedelta(minutes=30)
        row_index += 1


# Converts Room to correct format
# str -> str
def format_room_number(room_str):
    # Split the string by spaces
    parts = room_str.split()
    print(parts)
    # Find the part that contains the room number
    for part in parts:
        if '-' in part:
            # Remove leading zeros and return the formatted string
            formatted = part.replace('-', ' ').lstrip('0').replace(' ', '-')
            return formatted


# Takes a csv file and returns a list of directories
# file -> Array of Dictionaries
def csv_to_dicts(input_file):
    # Read data from CSV file
    df = pd.read_csv(input_file)

    # Remove all empty ROWS
    df = df.dropna(how='all')

    # Convert DataFrame to list of dictionaries
    data = df.to_dict(orient='records')
    return data


# Makes a timesheet from 7:00 AM to 10:00 PM
# void -> DataFrame (Excel Sheet)
def create_time_sheet(unique_rooms):
    # Initialize start time and end time
    start_time = datetime.strptime("07:00 AM", "%I:%M %p")
    end_time = datetime.strptime("10:00 PM", "%I:%M %p")
    
    # Create a list to hold the time values
    times = []
    
    # Generate times in 30-minute increments
    current_time = start_time
    while current_time <= end_time:
        times.append(current_time.strftime("%I:%M %p"))
        current_time += timedelta(minutes=30)
    
    # Create DataFrame
    df = pd.DataFrame(index=times)
    
    return df


def calculate_chunks(start_time, end_time):

    # Convert times to datetime objects
    time_format = "%I:%M %p"
    start = datetime.strptime(start_time, time_format)
    end = datetime.strptime(end_time, time_format)

    # Calculate the difference in minutes
    delta_minutes = int((end - start).total_seconds() / 60)

    # Calculate the number of 30-minute chunks
    return delta_minutes // 30

'''
MAIN FUNCTION THAT SCHEDULES LABS FOR 'MWF' AND 'TR':
sheet - ws
unique_rooms - list
lab_room - str
start_time - str
instructor - str
course - str
section - int
num_cells - int
'''
# Fills the appropriate cell with the proper format given the series of inputs
# in the 'XX Labs' Sheet
def schedule_lab(sheet, unique_rooms, lab_room, start_time, end_time, instructor, course, section, num_cells, days):    
    # Find the starting row based on the start time
    start_row = None
    for row in range(2, sheet.max_row + 1):
        time_value = sheet.cell(row=row, column=1).value
        if time_value == start_time:
            start_row = row
            break

    if start_row is None:
        print(f"Error: Start time {start_time} not found in the 'Time' column.")
        return
    
        # Determine the column based on the room
    try:
        column = unique_rooms.index(lab_room) + 2  # Assuming the first column is 'Time'
    except ValueError:
        print(f"Error: Room {lab_room} not found in unique_rooms.")
        return

    # Populate the cells with the lab class information
    formatted_value = f"{instructor}\n{course}-{int(section)}\nLaboratory\n{days} {start_time}-{end_time}"

    #for i in range(num_cells):
    cell = sheet.cell(row=start_row, column=column)
    cell.value = formatted_value
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    remake.merge_row_cells(sheet, start_row, num_cells, start_time, column)

# Fills the appropriate cell with the proper format given the series of inputs
# in the 'Instructor' Sheet
def schedule_instructor(sheet, unique_instructors, lab_room, start_time, end_time, instructor, course, section, num_cells, component, day):
    # Find the starting row based on the start time
    start_row = None
    for row in range(2, sheet.max_row + 1):
        time_value = sheet.cell(row=row, column=1).value
        if time_value == start_time:
            start_row = row
            break

    if start_row is None:
        print(f"Error: Start time {start_time} not found in the 'Time' column.")
        return
    
        # Determine the column based on the room
    try:
        column = unique_instructors.index(instructor) + 2  # Assuming the first column is 'Time'
    except ValueError:
        print(f"Error: Room {instructor} not found in unique_instructors.")
        return

    # Populate the cells with the lab class information
    if component == 'Lecture':
        formatted_value = f"{instructor}\n{course}-{int(section)}\n{component}\n{day} {start_time}-{end_time}"
    else:
        formatted_value = f"{instructor}\n{course}-{int(section)}\n{component}\n{day} {start_time}-{end_time}\nRoom {lab_room}"

    #for i in range(num_cells):
    cell = sheet.cell(row=start_row, column=column)
    cell.value = formatted_value
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # NEEDS TO BE DONE RECURSIVELY?
    #sheet.merge_cells(start_row=start_row, start_column=column, end_row=start_row+num_cells, end_column=column)

if __name__ == "__main__":
    main()